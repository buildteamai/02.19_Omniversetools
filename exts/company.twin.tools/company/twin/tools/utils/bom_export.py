# SPDX-FileCopyrightText: Copyright (c) 2024-2026 BuildTeam AI. All rights reserved.
# SPDX-License-Identifier: Proprietary

"""
USD Bill of Materials (BOM) Exporter

Extracts engineering metadata from USD stage prims and exports to Excel.
"""

from dataclasses import dataclass, field
from typing import Dict, Any, List, Optional
from collections import defaultdict
import json
import os
from pxr import Usd, UsdGeom

@dataclass
class BOMItem:
    """Represents a single BOM line item."""
    prim_path: str
    generator_type: str
    designation: str = ""
    description: str = ""
    length: float = 0.0
    quantity: int = 1
    weight_per_unit: float = 0.0
    total_weight: float = 0.0
    material: str = ""
    gauge: int = 0
    dimensions: str = ""
    custom_data: Dict[str, Any] = field(default_factory=dict)
    
    def __post_init__(self):
        if self.weight_per_unit > 0 and self.length > 0:
            # Weight is typically per foot, convert length from inches
            self.total_weight = self.weight_per_unit * (self.length / 12.0) * self.quantity


class BOMExporter:
    """
    Extracts Bill of Materials data from USD stage prims.
    """
    
    # Generator type configurations
    GENERATOR_CONFIGS = {
        'wide_flange': {
            'category': 'Steel',
            'description_template': 'Wide Flange Beam',
            'weight_key': 'weight_lb_ft',
            'designation_key': 'designation',
        },
        'channel': {
            'category': 'Steel',
            'description_template': 'C-Channel',
            'weight_key': 'weight_lb_ft',
            'designation_key': 'designation',
        },
        'hss_tube': {
            'category': 'Steel',
            'description_template': 'HSS Tube',
            'weight_key': 'weight_lb_ft',
            'designation_key': 'designation',
        },
        'sheet_metal_panel': {
            'category': 'Sheet Metal',
            'description_template': 'Sheet Metal Panel',
        },
        'sheet_metal_subpart': {
            'category': 'Sheet Metal - Parts',
            'description_template': 'Part',
        },
        'glazing_panel': {
            'category': 'Glazing',
            'description_template': 'Glazing',
        },
        'duct_straight': {
            'category': 'MEP - Ductwork',
            'description_template': 'Rectangular Duct - Straight',
        },
        'duct_bent': {
            'category': 'MEP - Ductwork',
            'description_template': 'Rectangular Duct - Elbow',
        },
        'duct_round_straight': {
            'category': 'MEP - Ductwork',
            'description_template': 'Round Duct - Straight',
        },
        'duct_round_bent': {
            'category': 'MEP - Ductwork',
            'description_template': 'Round Duct - Elbow',
        },
        'pipe': {
            'category': 'MEP - Piping',
            'description_template': 'Pipe',
        },
        'pipe_straight': {
            'category': 'MEP - Piping',
            'description_template': 'Pipe - Straight',
        },
        'pipe_bent': {
            'category': 'MEP - Piping',
            'description_template': 'Pipe - Elbow',
        },
        'pyramid': {
            'category': 'Objects',
            'description_template': 'Pyramid',
        },
        'gusset_plate': {
            'category': 'Steel Connections',
            'description_template': 'Gusset Plate',
        },
        'shear_tab': {
            'category': 'Steel Connections',
            'description_template': 'Shear Tab',
        },
        'double_angle': {
            'category': 'Steel Connections',
            'description_template': 'Double Angle',
        },
        'equipment': {
            'category': 'Equipment',
            'description_template': 'Equipment',
        },
        'inline_component': {
            'category': 'MEP - Components',
            'description_template': 'Component',
        },
        'door_slab': {
            'category': 'Doors',
            'description_template': 'Door Slab',
        },
        # Generic Types
        'generic_mesh': {
            'category': 'Scene - Geometry',
            'description_template': 'Mesh',
        },
        'generic_xform': {
            'category': 'Scene - Transforms',
            'description_template': 'Group/Xform',
        },
        'generic_light': {
            'category': 'Scene - Lights',
            'description_template': 'Light',
        },
        'generic_camera': {
            'category': 'Scene - Cameras',
            'description_template': 'Camera',
        },
        'generic_scope': {
            'category': 'Scene - Scopes',
            'description_template': 'Scope',
        },
        'generic_prim': {
            'category': 'Scene - Other',
            'description_template': 'Object',
        },
    }
    
    @staticmethod
    def extract_from_stage(stage) -> List[BOMItem]:
        """
        Traverses USD stage and extracts BOM data from ALL prims.
        
        Args:
            stage: USD Stage object
            
        Returns:
            List of BOMItem objects
        """
        items = []
        
        if not stage:
            print("[BOMExporter] Error: No stage provided")
            return items
        
        # Traverse all prims, including instance proxies for deep nested assets
        # We need to traverse in a way that lets us skip children if parent is a BOM item
        # Usd.Stage.Traverse is depth-first Pre-Order. 
        # Ideally we want to identify a "Generator" (Component) and ignore its internal geometry.
        
        processed_paths = set()
        
        for prim in stage.Traverse(Usd.TraverseInstanceProxies()):
            path = str(prim.GetPath())
            
            try:
                # 1. HIERARCHY CHECK
                # If this prim is a child of something we already processed/added, SKIP IT.
                # However, if the child is EXPLICITLY tagged as a BOM item (like a sub-part), we ALLOW it.
                is_child = False
                for parent_path in processed_paths:
                    if path.startswith(parent_path + "/"):
                        is_child = True
                        break
                
                # Check if we should override the skip
                # We need the generator type to decide. 
                custom_data = prim.GetCustomData() or {}
                generator_type = custom_data.get('generatorType')
                if not generator_type:
                    generator_type = BOMExporter._infer_generator_type(prim)
                    
                allowed_nested = ['sheet_metal_subpart', 'glazing_panel', 'door_slab']
                
                if is_child:
                    if generator_type and generator_type in allowed_nested:
                        # Allow it!
                        pass
                    else:
                        continue
                
                # EXCLUSION LOGIC
                # Strict filtering: If it returns None or specific ignored types, skip.
                if not generator_type or generator_type in ['generic_ignored']:
                    continue
                
                # Extract BOM item from prim
                item = BOMExporter._extract_item_from_prim(prim, custom_data, generator_type)
                if item:
                    items.append(item)
                    # Mark this path as processed so we skip its children
                    processed_paths.add(path)
            
            except Exception as e:
                print(f"[BOMExporter] Error processing prim {path}: {e}")
                continue
        
        print(f"[BOMExporter] Extracted {len(items)} items from stage")
        print(f"[BOMExporter] Extracted {len(items)} items from stage")
        return items

    @staticmethod
    def extract_frames(stage) -> List[Dict[str, Any]]:
        """
        Extracts structural frame data for engineering reports.
        """
        frames = []
        if not stage: return frames
        
        for prim in stage.Traverse():
            try:
                if prim.GetCustomDataByKey("generatorType") == "structural_frame":
                    eng_data_str = prim.GetCustomDataByKey("engineering_data")
                    if eng_data_str:
                        eng_data = json.loads(eng_data_str)
                        frames.append({
                            'path': str(prim.GetPath()),
                            'name': prim.GetName(),
                            'data': eng_data
                        })
            except Exception:
                continue
        return frames

    @staticmethod
    def _infer_generator_type(prim) -> Optional[str]:
        """Infers a generic generator type from the USD Prim type."""
        
        # 1. Check for legacy Sheet Metal Panels (by attribute)
        if prim.HasAttribute("custom:panel_type"):
            return 'sheet_metal_panel'
            
        # 2. Check for legacy Sheet Metal Panels (by name) - fallback
        if "SheetMetalPanel" in prim.GetName():
            return 'sheet_metal_panel'

        if prim.IsA(UsdGeom.Mesh):
            # STRICT FILTERING:
            # The fabricator doesn't want random "Scene Meshes".
            # Only include meshes if they have explicit metadata or are part of a known system.
            # If it's just a raw mesh without generatorType, IGNORE IT.
            return None
            
        # IGNORE EVERYTHING ELSE
        # We explicitly do NOT want generic_xform, lights, cameras, scopes, etc.
        # changing return to 'generic_ignored' or None
        
        return None
    
    @staticmethod
    def _extract_item_from_prim(prim, custom_data: Dict, generator_type: str) -> Optional[BOMItem]:
        """Extracts BOM data from a single prim."""
        config = BOMExporter.GENERATOR_CONFIGS.get(generator_type, {})
        
        # Get designation
        designation_key = config.get('designation_key', 'designation')
        designation = custom_data.get(designation_key, '')
        
        # Get length
        length = custom_data.get('length', 0.0)
        if isinstance(length, str):
            try:
                length = float(length)
            except ValueError:
                length = 0.0
        
        # Get weight per unit
        weight_per_unit = 0.0
        weight_key = config.get('weight_key')
        if weight_key:
            # Weight might be in aisc_data
            aisc_data = custom_data.get('aisc_data')
            if aisc_data:
                if isinstance(aisc_data, str):
                    try:
                        aisc_data = json.loads(aisc_data)
                    except json.JSONDecodeError:
                        aisc_data = {}
                weight_per_unit = aisc_data.get(weight_key, 0.0)
        
        # Build dimensions string
        dimensions = BOMExporter._build_dimensions_string(custom_data, generator_type)
        
        # Build description
        description = config.get('description_template', generator_type.replace('_', ' ').title())
        if designation:
            description = f"{description} - {designation}"
        
        # Get material
        material = custom_data.get('pipeMaterial', custom_data.get('material', ''))
        if generator_type in ['wide_flange', 'channel', 'hss_tube']:
            material = 'A992 Steel' if not material else material
            
        # Get gauge
        gauge = custom_data.get('gauge', 0)

        # Calculate Weight
        total_weight = 0.0
        
        # SHEET METAL WEIGHT CALCULATION
        if generator_type in ['sheet_metal_panel', 'sheet_metal_subpart', 'door_slab', 'gusset_plate', 'shear_tab', 'double_angle']:
            # Weight = Volume * Density
            # Steel Density ~ 0.2836 lbs/in^3
            w = float(custom_data.get('width', 0))
            h = float(custom_data.get('height', 0))
            t = float(custom_data.get('thickness', 0))
            
            # If thickness missing (some old parts might use gauge lookup?)
            # For now rely on 'thickness' attribute being present
            if w > 0 and h > 0 and t > 0:
                volume = w * h * t
                density = 0.2836 # lbs/in^3
                total_weight = volume * density
                
                # Approximate flanges for main panels (adds ~10-15% surface area usually)
                # Or use explicit flange depth if available? 
                # Keeping it simple: Flat Area * Thickness * Density
                # PENDING: Add flange area if strictly needed.
        
        
        return BOMItem(
            prim_path=str(prim.GetPath()),
            generator_type=generator_type,
            designation=designation,
            description=description,
            length=length,
            quantity=1,
            weight_per_unit=weight_per_unit,
            total_weight=total_weight,
            material=material,
            gauge=gauge,
            dimensions=dimensions,
            custom_data=custom_data
        )
    
    @staticmethod
    def _build_dimensions_string(custom_data: Dict, generator_type: str) -> str:
        """Builds a human-readable dimensions string."""
        parts = []
        
        if generator_type in ['duct_straight', 'duct_bent']:
            w = custom_data.get('width', 0)
            h = custom_data.get('height', 0)
            if w and h:
                parts.append(f"{w}\"W x {h}\"H")
        
        elif generator_type in ['duct_round_straight', 'duct_round_bent']:
            d = custom_data.get('diameter', 0)
            if d:
                parts.append(f"{d}\" Dia")
        
        elif generator_type == 'pipe':
            d = custom_data.get('diameter', 0)
            if d:
                parts.append(f"{d}\" NPS")
        
        elif generator_type == 'pyramid':
            base = custom_data.get('base', 0)
            height = custom_data.get('height', 0)
            if base and height:
                parts.append(f"Base: {base}\", Height: {height}\"")
        
        length = custom_data.get('length', 0)
        if length and generator_type not in ['pyramid']:
            parts.append(f"L: {length}\"")
            
        # Add Gauge if present
        gauge = custom_data.get('gauge')
        if gauge:
            parts.append(f"{gauge}ga")
        
        return ', '.join(parts)
    
    @staticmethod
    def rollup_bom(items: List[BOMItem]) -> List[BOMItem]:
        """
        Groups BOM items by type and designation, summing quantities.
        
        Args:
            items: List of individual BOM items
            
        Returns:
            List of rolled-up BOM items with aggregated quantities
        """
        rollup = defaultdict(lambda: {
            'quantity': 0,
            'total_length': 0.0,
            'total_weight': 0.0,
            'paths': [],
            'item': None
        })
        
        for item in items:
            # Group by type, designation, dimensions, MATERIAL, and GAUGE
            key = (item.generator_type, item.designation, item.dimensions, item.material, item.gauge)
            rollup[key]['quantity'] += item.quantity
            rollup[key]['total_length'] += item.length
            rollup[key]['total_weight'] += item.total_weight
            rollup[key]['paths'].append(item.prim_path)
            
            if rollup[key]['item'] is None:
                rollup[key]['item'] = item
        
        result = []
        for key, data in rollup.items():
            item = data['item']
            rolled_item = BOMItem(
                prim_path=', '.join(data['paths'][:3]) + ('...' if len(data['paths']) > 3 else ''),
                generator_type=item.generator_type,
                designation=item.designation,
                description=item.description,
                length=data['total_length'],
                quantity=data['quantity'],
                weight_per_unit=item.weight_per_unit,
                total_weight=data['total_weight'],
                material=item.material,
                gauge=item.gauge,
                dimensions=item.dimensions,
                custom_data={}
            )
            result.append(rolled_item)
        
        # Sort by category, then generator type, then designation
        result.sort(key=lambda x: (
            BOMExporter.GENERATOR_CONFIGS.get(x.generator_type, {}).get('category', 'Other'),
            x.generator_type,
        x.designation
        ))
        
        return result
    
    @staticmethod
    def export_to_excel(items: List[BOMItem], filepath: str, project_name: str = "Project", stage=None) -> bool:
        """
        Exports BOM items to an Excel spreadsheet.
        Optionally exports Engineering Data if stage is provided to extract frames.
        
        Args:
            items: List of BOM items to export
            filepath: Output file path (.xlsx)
            project_name: Name of project for header
            stage: Optional. The USD stage to extract structural frame data from.
            
        Returns:
            True if export successful, False otherwise
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("[BOMExporter] openpyxl not available, falling back to CSV")
            return BOMExporter.export_to_csv(items, filepath.replace('.xlsx', '.csv'))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Bill of Materials"
        
        # Styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True, size=16)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title row
        ws.merge_cells('A1:H1')
        ws['A1'] = f"Bill of Materials - {project_name}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Item #', 'Type', 'Description', 'Dimensions', 'Qty', 'Unit Wt (lb/ft)', 'Total Wt (lb)', 'Material']
        header_row = 3
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data rows
        for idx, item in enumerate(items, 1):
            row = header_row + idx
            
            category = BOMExporter.GENERATOR_CONFIGS.get(item.generator_type, {}).get('category', 'Other')
            
            row_data = [
                idx,
                category,
                item.description,
                item.dimensions,
                item.quantity,
                f"{item.weight_per_unit:.2f}" if item.weight_per_unit > 0 else "-",
                f"{item.total_weight:.2f}" if item.total_weight > 0 else "-",
                item.material
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col in [5, 6, 7]:  # Numeric columns
                    cell.alignment = Alignment(horizontal='right')
        
        # Summary row
        if items:
            summary_row = header_row + len(items) + 2
            ws.cell(row=summary_row, column=1, value="TOTALS:").font = Font(bold=True)
            ws.cell(row=summary_row, column=5, value=sum(i.quantity for i in items)).font = Font(bold=True)
            total_weight = sum(i.total_weight for i in items)
            ws.cell(row=summary_row, column=7, value=f"{total_weight:.2f}").font = Font(bold=True)
        
        # Auto-size columns
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
            
        # ---------------------------------------------------------
        # SHEET 2: STRUCTURAL CALCS
        # ---------------------------------------------------------
        if stage:
            frames = BOMExporter.extract_frames(stage)
            if frames:
                ws_calcs = wb.create_sheet("Structural Calcs")
                
                # Title
                ws_calcs.merge_cells('A1:F1')
                ws_calcs['A1'] = "Engineering Analysis Report"
                ws_calcs['A1'].font = title_font
                ws_calcs['A1'].alignment = Alignment(horizontal='center')
                
                # Headers
                calc_headers = ['Frame ID', 'Load (lbs)', 'Deflection (in)', 'Limit (in)', 'Stress Ratio', 'Status']
                h_row = 3
                
                for col, header in enumerate(calc_headers, 1):
                    cell = ws_calcs.cell(row=h_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                    
                # Data
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Light Green
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Light Red
                
                for idx, frame in enumerate(frames, 1):
                    r = h_row + idx
                    data = frame['data']
                    
                    # Stress Ratio = Stress / Limit
                    stress = data.get('stress', 0)
                    limit_stress = data.get('limit_stress', 1)
                    ratio = stress / limit_stress if limit_stress > 0 else 0
                    
                    row_vals = [
                        frame['name'],
                        data.get('point_load_lbs', 0),
                        f"{data.get('deflection', 0):.4f}",
                        f"{data.get('limit_deflection', 0):.4f}",
                        f"{ratio:.2%}",
                        data.get('status', 'UNKNOWN')
                    ]
                    
                    status = data.get('status', 'FAIL')
                    row_fill = green_fill if status == 'PASS' else red_fill
                    
                    for col, val in enumerate(row_vals, 1):
                        cell = ws_calcs.cell(row=r, column=col, value=val)
                        cell.border = thin_border
                        cell.fill = row_fill
                        if col in [2,3,4,5]: cell.alignment = Alignment(horizontal='right')

                # Column Widths
                calc_widths = [20, 15, 15, 15, 15, 10]
                for col, width in enumerate(calc_widths, 1):
                    ws_calcs.column_dimensions[get_column_letter(col)].width = width

        # Save
        wb.save(filepath)
        print(f"[BOMExporter] Exported {len(items)} items to {filepath}")
        return True
    
    @staticmethod
    def export_to_csv(items: List[BOMItem], filepath: str) -> bool:
        """
        Fallback: Exports BOM to CSV file.
        
        Args:
            items: List of BOM items
            filepath: Output file path (.csv)
            
        Returns:
            True if successful
        """
        import csv
        
        headers = ['Item #', 'Type', 'Description', 'Dimensions', 'Qty', 'Unit Wt (lb/ft)', 'Total Wt (lb)', 'Material']
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            
            for idx, item in enumerate(items, 1):
                category = BOMExporter.GENERATOR_CONFIGS.get(item.generator_type, {}).get('category', 'Other')
                row = [
                    idx,
                    category,
                    item.description,
                    item.dimensions,
                    item.quantity,
                    f"{item.weight_per_unit:.2f}" if item.weight_per_unit > 0 else "",
                    f"{item.total_weight:.2f}" if item.total_weight > 0 else "",
                    item.material
                ]
                writer.writerow(row)
        
        print(f"[BOMExporter] Exported {len(items)} items to CSV: {filepath}")
        return True
